#!/usr/bin/env python3
"""
Fetch Grinn model scenario snapshots by downloading the .xlsm file
via Microsoft Graph API and parsing it locally with openpyxl.

This avoids Excel Online's calculation engine (which 504s on complex workbooks)
by downloading the raw file and reading cell values directly.
"""
import json, os, sys, tempfile
from datetime import datetime, timezone
from urllib.request import Request, urlopen
from urllib.error import HTTPError

# ── Config ──────────────────────────────────────────────────────────────
TENANT_ID     = os.environ["AZURE_TENANT_ID"]
CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
USER_ID       = os.environ["GRAPH_USER_ID"]
DRIVE_PATH    = os.environ.get("DRIVE_PATH", "Grinn/Grinn_v17_r41.xlsm")
SHEET         = os.environ.get("SHEET", "45_Combined")

GRAPH_BASE    = "https://graph.microsoft.com/v1.0"
TOKEN_URL     = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"

SCENARIOS     = ["Bull", "Base", "Bear", "Organic"]
SNAPSHOT_STARTS = {"Bull": 30, "Base": 59, "Bear": 88, "Organic": 114}
N_ROWS        = 24
COL_COUNT     = 6  # A through F
YEARS         = ["2026", "2027", "2028", "2029", "2030"]

# ── Auth ────────────────────────────────────────────────────────────────
def get_app_token():
    body = (
        f"client_id={CLIENT_ID}"
        f"&client_secret={CLIENT_SECRET}"
        f"&scope=https%3A%2F%2Fgraph.microsoft.com%2F.default"
        f"&grant_type=client_credentials"
    ).encode()
    req = Request(TOKEN_URL, data=body, headers={"Content-Type": "application/x-www-form-urlencoded"})
    resp = json.loads(urlopen(req).read())
    return resp["access_token"]

# ── Download ────────────────────────────────────────────────────────────
def download_workbook(token):
    """Download the raw .xlsm file via Graph API (no Excel Online needed)."""
    url = f"{GRAPH_BASE}/users/{USER_ID}/drive/root:/{DRIVE_PATH}:/content"
    req = Request(url, headers={"Authorization": f"Bearer {token}"})

    # Follow redirects - Graph API returns a 302 to the actual download URL
    import urllib.request
    opener = urllib.request.build_opener(urllib.request.HTTPRedirectHandler)
    resp = opener.open(req, timeout=300)
    return resp.read()

# ── Main ────────────────────────────────────────────────────────────────
def main():
    # Step 0: Install openpyxl
    print("📦 Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

    print("🔐 Authenticating (client credentials)...")
    token = get_app_token()
    print("✅ Token acquired")

    # Step 1: Download the raw workbook file
    print(f"📥 Downloading {DRIVE_PATH}...")
    file_bytes = download_workbook(token)
    print(f"  Downloaded {len(file_bytes):,} bytes")

    # Step 2: Save to temp file and open with openpyxl
    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    print(f"📊 Opening workbook with openpyxl...")
    import openpyxl
    # data_only=True reads cached values instead of formulas
    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True, keep_links=False)

    if SHEET not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET}' not found! Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET]
    print(f"  Sheet '{SHEET}' loaded")

    # Step 3: Read each scenario snapshot
    output = {
        "updated": datetime.now(timezone.utc).isoformat(),
        "sheet": SHEET,
        "years": YEARS,
        "scenarios": {}
    }

    for scenario in SCENARIOS:
        start_row = SNAPSHOT_STARTS[scenario]
        print(f"📊 Reading {scenario} (rows {start_row}-{start_row + N_ROWS - 1})...")

        row_labels = []
        scenario_data = {}

        for yi, year in enumerate(YEARS):
            year_data = {}
            for offset in range(N_ROWS):
                row_num = start_row + offset
                col_a = ws.cell(row=row_num, column=1).value  # Label column
                col_val = ws.cell(row=row_num, column=yi + 2).value  # Year columns B-F

                label = str(col_a or "").strip() if col_a else f"row_{offset}"
                if yi == 0:
                    row_labels.append(label)

                # Convert to number if possible, default to 0
                if col_val is None:
                    col_val = 0
                elif isinstance(col_val, str):
                    try:
                        col_val = float(col_val)
                    except ValueError:
                        col_val = 0

                year_data[label] = col_val
            scenario_data[year] = year_data

        output["scenarios"][scenario] = {
            "labels": row_labels,
            "data": scenario_data
        }
        print(f"  ✅ {scenario}: {len(row_labels)} rows, labels: {row_labels[:5]}...")

    wb.close()
    os.unlink(tmp_path)

    if not output["scenarios"]:
        print("❌ No scenarios were read successfully!")
        sys.exit(1)

    # Write output
    out_path = os.environ.get("OUTPUT_PATH", "data.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2)
    print(f"\n📁 Wrote {out_path} ({os.path.getsize(out_path)} bytes)")
    print(f"🕐 Updated: {output['updated']}")

    # Print a sample for verification
    first_scenario = SCENARIOS[0]
    first_year = YEARS[0]
    sample = output["scenarios"][first_scenario]["data"][first_year]
    print(f"\n📋 Sample ({first_scenario}/{first_year}):")
    for label, val in list(sample.items())[:5]:
        print(f"  {label}: {val}")

if __name__ == "__main__":
    main()
